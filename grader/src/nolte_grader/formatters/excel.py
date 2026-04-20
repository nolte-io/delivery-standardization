"""Excel workbook formatter — two-sheet output for auditability.

Sheet 1 "Rollup": summary stats (pass rate, dimension fail rates, cycle time,
    owner breakdown, recommendations).
Sheet 2 "Detail": one row per issue with a clickable Jira URL, all dimension
    verdicts, and key timestamps.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from nolte_grader.core.models import IssueGrade, RollupReport, Verdict
from nolte_grader.formatters.markdown import _DIM_LABELS, _DIM_ACTION, _build_recommendations

# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------

_GREEN = PatternFill("solid", fgColor="C6EFCE")
_RED = PatternFill("solid", fgColor="FFC7CE")
_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_GREY = PatternFill("solid", fgColor="D9D9D9")
_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)

_VERDICT_FILL: dict[str, PatternFill] = {
    Verdict.PASS: _GREEN,
    Verdict.FAIL: _RED,
    Verdict.INSUFFICIENT_EVIDENCE: _YELLOW,
    Verdict.NOT_APPLICABLE: _GREY,
}

_DIM_CODES = [
    "Y1", "Y2", "Y3", "Y4", "Y5", "Y6",
    "U7", "U8", "U9", "U10", "U11", "U12",
    "C1", "C2", "C3",
    "D1", "D2", "D3", "D4", "D5", "D6", "D7", "D8", "D9", "D10",
]


def write_workbook(
    report: RollupReport,
    grades: list[IssueGrade],
    instance_url: str,
    path: Path,
) -> None:
    """Write grades.xlsx to *path* with Rollup and Detail sheets."""
    wb = openpyxl.Workbook()
    ws_rollup = wb.active
    ws_rollup.title = "Rollup"
    ws_detail = wb.create_sheet("Detail")

    _build_rollup_sheet(ws_rollup, report)
    _build_detail_sheet(ws_detail, grades, instance_url)

    wb.save(path)


# ---------------------------------------------------------------------------
# Sheet 1 — Rollup
# ---------------------------------------------------------------------------

def _build_rollup_sheet(ws: Any, report: RollupReport) -> None:
    sys = report.system

    def _header(row: int, text: str) -> None:
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=12)

    def _kv(row: int, key: str, value: Any) -> None:
        ws.cell(row=row, column=1, value=key).font = _BOLD
        ws.cell(row=row, column=2, value=value)

    r = 1
    ws.cell(row=r, column=1, value="Delivery Report").font = Font(bold=True, size=14)
    r += 1
    ws.cell(row=r, column=1,
            value=f"{sys.window.from_date} → {sys.window.to_date}")
    r += 2

    _header(r, "Summary"); r += 1
    _kv(r, "Run ID", sys.run_id); r += 1
    _kv(r, "Items graded", sys.issue_count_all_types); r += 1
    _kv(r, "Six-yes pass rate", f"{sys.six_yes_pass_rate:.0%}"); r += 1
    _kv(r, "Pending judge dimensions", sys.pending_judge_dimensions); r += 2

    _header(r, "Cycle Time (Done Specifying → Done Implementing)"); r += 1
    _kv(r, "p50 (days)", round(sys.cycle_time_p50, 2) if sys.cycle_time_p50 is not None else "—"); r += 1
    _kv(r, "p90 (days)", round(sys.cycle_time_p90, 2) if sys.cycle_time_p90 is not None else "—"); r += 1
    _kv(r, "max (days)", round(sys.cycle_time_max, 2) if sys.cycle_time_max is not None else "—"); r += 1
    _kv(r, f"Above {sys.cycle_time_threshold_days}d threshold",
        sys.cycle_time_stories_above_threshold); r += 2

    _header(r, "Dimension Failures"); r += 1
    col_headers = ["Code", "Dimension", "Fail rate", "Failed", "Graded"]
    for ci, h in enumerate(col_headers, 1):
        cell = ws.cell(row=r, column=ci, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    r += 1

    sorted_dims = sorted(
        sys.dimension_fail_rates.values(),
        key=lambda d: (-d.fail_rate, d.code),
    )
    for dfr in sorted_dims:
        if dfr.graded == 0:
            continue
        pct = f"{dfr.fail_rate:.0%}"
        for ci, val in enumerate(
            [dfr.code, _DIM_LABELS.get(dfr.code, dfr.code), pct, dfr.fails, dfr.graded], 1
        ):
            ws.cell(row=r, column=ci, value=val)
        fill = _RED if dfr.fail_rate >= 0.5 else (_YELLOW if dfr.fail_rate > 0 else _GREEN)
        ws.cell(row=r, column=3).fill = fill
        r += 1
    r += 1

    _header(r, "Owner Breakdown — Engineers"); r += 1
    for ci, h in enumerate(["Engineer", "Items", "Top failures"], 1):
        cell = ws.cell(row=r, column=ci, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    r += 1
    for eng, stats in sorted(report.owner.downstream_by_engineer.items()):
        ws.cell(row=r, column=1, value=eng)
        ws.cell(row=r, column=2, value=stats.stories)
        ws.cell(row=r, column=3, value=", ".join(stats.top_fails))
        r += 1
    r += 1

    _header(r, "Recommendations"); r += 1
    for i, rec in enumerate(_build_recommendations(report), 1):
        ws.cell(row=r, column=1, value=f"{i}.")
        cell = ws.cell(row=r, column=2, value=rec)
        cell.alignment = Alignment(wrap_text=True)
        r += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10


# ---------------------------------------------------------------------------
# Sheet 2 — Detail
# ---------------------------------------------------------------------------

def _build_detail_sheet(ws: Any, grades: list[IssueGrade], instance_url: str) -> None:
    base_url = instance_url.rstrip("/")

    fixed_cols = [
        "Issue Key", "Jira Link", "Type", "Project", "Epic",
        "Title",
        "Commitment Date", "Done Date", "Cycle Time (days)",
        "Six-Yes", "Upstream", "Downstream", "Story Overall",
        "Spec Approver", "Downstream Owner",
    ]
    dim_verdict_cols = [f"{c} verdict" for c in _DIM_CODES]
    dim_evidence_cols = [f"{C} evidence" for C in _DIM_CODES]

    all_headers = fixed_cols + dim_verdict_cols + dim_evidence_cols

    for ci, h in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    for ri, g in enumerate(grades, 2):
        url = f"{base_url}/browse/{g.issue_key}"
        row: list[Any] = [
            g.issue_key,
            f'=HYPERLINK("{url}", "{g.issue_key}")',
            g.issue_type,
            g.project_key,
            g.epic_key or "",
            g.title,
            g.commitment_timestamp.date().isoformat() if g.commitment_timestamp else "",
            g.done_timestamp.date().isoformat() if g.done_timestamp else "",
            round(g.cycle_time_days, 2) if g.cycle_time_days is not None else "",
            g.six_yes_overall,
            g.upstream_overall,
            g.downstream_overall,
            g.story_overall,
            g.spec_approver or "",
            g.downstream_owner or "",
        ]
        for code in _DIM_CODES:
            dim = g.dimensions.get(code)
            row.append(dim.verdict if dim else "")
        for code in _DIM_CODES:
            dim = g.dimensions.get(code)
            row.append(dim.evidence_code if dim else "")

        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

        # Colour overall verdict columns
        overall_col_indices = {
            "Six-Yes": fixed_cols.index("Six-Yes") + 1,
            "Upstream": fixed_cols.index("Upstream") + 1,
            "Downstream": fixed_cols.index("Downstream") + 1,
            "Story Overall": fixed_cols.index("Story Overall") + 1,
        }
        for col_name, col_idx in overall_col_indices.items():
            verdict_val = ws.cell(row=ri, column=col_idx).value
            fill = _VERDICT_FILL.get(str(verdict_val), _GREY)
            ws.cell(row=ri, column=col_idx).fill = fill

        # Colour per-dim verdict cells
        verdict_start = len(fixed_cols) + 1
        for di, code in enumerate(_DIM_CODES):
            col_idx = verdict_start + di
            verdict_val = ws.cell(row=ri, column=col_idx).value
            fill = _VERDICT_FILL.get(str(verdict_val), _GREY)
            ws.cell(row=ri, column=col_idx).fill = fill

    # Freeze header row and set column widths
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 16
    for col_idx in range(10, len(all_headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
